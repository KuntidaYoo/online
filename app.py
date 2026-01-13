import datetime as dt
import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple, Any
from io import BytesIO

import streamlit as st
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from copy import copy as _copy

from pathlib import Path


# ============================================================
# Core logic (same as your script) — safe for Streamlit deploy
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "template_output.xlsx"

def thai_be_to_yy(date_greg: dt.date) -> str:
    """Thai BE year last-2 digits. 2026 -> 2569 -> '69'."""
    be_year = date_greg.year + 543
    return str(be_year % 100).zfill(2)

def build_po_base(branch_code: str, po_date: dt.date) -> tuple[str, str]:
    """
    Returns (base, num_no_zero)
    base = YYMMDD + num_no_zero
    """
    m = re.match(r"^(SPX|TT|LEX)(\d+)$", branch_code.strip().upper())
    if not m:
        raise ValueError(f"Invalid branch code: {branch_code} (e.g., SPX903, LEX905, TT905)")

    num = m.group(2)
    yy = thai_be_to_yy(po_date)
    mmdd = po_date.strftime("%m%d")

    num_no_zero = num.replace("0", "") or "0"
    base = f"{yy}{mmdd}{num_no_zero}"
    return base, num_no_zero

def copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int = 18) -> None:
    """Copy style from template row to new rows to preserve formatting."""
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        dst._style = _copy(src._style)
        dst.number_format = src.number_format
        dst.font = _copy(src.font)
        dst.border = _copy(src.border)
        dst.fill = _copy(src.fill)
        dst.alignment = _copy(src.alignment)
        dst.protection = _copy(src.protection)

def force_yyyy_mm_dd(date_obj: dt.date) -> str:
    """Always output yyyy/mm/dd format."""
    return date_obj.strftime("%Y/%m/%d")

def to_float(x: Any) -> float:
    """Convert Excel/CSV cell values to float safely."""
    if x is None or x == "":
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", "")
    try:
        return float(s)
    except Exception:
        return 0.0


@dataclass
class ParsedLine:
    description: str
    price: float
    qty: float


# ----------------------------
# Key lookup: key col B -> col C
# ----------------------------
def load_key_map_from_bytes(key_bytes: bytes, sheet_name: Optional[str] = None) -> Dict[str, str]:
    """Read key.xlsx from uploaded bytes, return mapping: B -> C."""
    wb = openpyxl.load_workbook(BytesIO(key_bytes), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    out: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if b is None or c is None:
            continue
        out[str(b).strip()] = str(c).strip()
    return out


# ----------------------------
# Parsers by fixed columns (consistent exports)
# ----------------------------
def parse_tt_bytes(xlsx_bytes: bytes) -> list[ParsedLine]:
    """
    TT:
      - starts row 3
      - G(7)=Seller SKU
      - J(10)=Quantity
      - L(12)=SKU Unit Original Price
    """
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    DESC_COL = 7
    QTY_COL = 10
    PRICE_COL = 12

    out: list[ParsedLine] = []
    for r in range(3, ws.max_row + 1):
        desc = ws.cell(r, DESC_COL).value
        if not desc:
            continue
        qty = to_float(ws.cell(r, QTY_COL).value)
        price = to_float(ws.cell(r, PRICE_COL).value)
        out.append(ParsedLine(str(desc).strip(), price, qty))
    return out


def parse_spx_bytes(xlsx_bytes: bytes) -> list[ParsedLine]:
    """
    SPX:
      - row 1 header, data from row 2
      - S(19)=SKU Reference No.
      - V(22)=ราคาขาย
      - W(23)=จำนวน
    """
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    DESC_COL = 19
    PRICE_COL = 22
    QTY_COL = 23

    out: list[ParsedLine] = []
    for r in range(2, ws.max_row + 1):
        desc = ws.cell(r, DESC_COL).value
        if not desc:
            continue
        price = to_float(ws.cell(r, PRICE_COL).value)
        qty = to_float(ws.cell(r, QTY_COL).value)
        out.append(ParsedLine(str(desc).strip(), price, qty))
    return out


def parse_lex_bytes(xlsx_bytes: bytes) -> list[ParsedLine]:
    """
    LEX:
      - row 1 header, data from row 2
      - F(6)=sellerSku
      - AV(48)=unitPrice
      - qty always 1
    """
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    DESC_COL = 6
    PRICE_COL = 48

    out: list[ParsedLine] = []
    for r in range(2, ws.max_row + 1):
        desc = ws.cell(r, DESC_COL).value
        if not desc:
            continue
        price = to_float(ws.cell(r, PRICE_COL).value)
        out.append(ParsedLine(str(desc).strip(), price, 1.0))
    return out


# ----------------------------
# Output rows and writing to template bytes
# ----------------------------
@dataclass
class OutputRow:
    branch_code: str
    article: str
    description: str
    price_unit: float
    qty: float


def build_output_rows_from_uploads(
    key_bytes: bytes,
    inputs: List[Tuple[str, str, bytes]],  # (platform, branch_code, file_bytes)
) -> List[OutputRow]:
    key_map = load_key_map_from_bytes(key_bytes)

    out: List[OutputRow] = []
    for platform, branch_code, file_bytes in inputs:
        platform = platform.upper().strip()
        branch_code = branch_code.strip().upper()

        if platform == "SPX":
            lines = parse_spx_bytes(file_bytes)
        elif platform == "LEX":
            lines = parse_lex_bytes(file_bytes)
        elif platform == "TT":
            lines = parse_tt_bytes(file_bytes)
        else:
            raise ValueError(f"Unknown platform: {platform}")

        for ln in lines:
            desc = ln.description
            article = key_map.get(desc, "")  # blank if not found
            out.append(OutputRow(
                branch_code=branch_code,
                article=article,
                description=desc,
                price_unit=ln.price,
                qty=(1.0 if platform == "LEX" else ln.qty),
            ))
    return out


def write_to_template_from_path(
    template_path: Path,
    rows: List[OutputRow],
    po_date: Optional[dt.date] = None,
    deliv_plus_days: int = 7,
) -> bytes:
    po_date = po_date or dt.date.today()
    po_str = force_yyyy_mm_dd(po_date)
    deliv_str = force_yyyy_mm_dd(po_date + dt.timedelta(days=deliv_plus_days))

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    start_row = 2
    po_by_branch: dict[str, str] = {}
    count_by_num: dict[str, int] = {}

    for i, rr in enumerate(rows):
        r = start_row + i
        if r != 2:
            copy_row_style(ws, src_row=2, dst_row=r, max_col=18)

        if rr.branch_code in po_by_branch:
            po_no = po_by_branch[rr.branch_code]
        else:
            base, num_no_zero = build_po_base(rr.branch_code, po_date)
            count_by_num[num_no_zero] = count_by_num.get(num_no_zero, 0) + 1
            suffix = str(count_by_num[num_no_zero]).zfill(2)
            po_no = base + suffix
            po_by_branch[rr.branch_code] = po_no

        ws[f"A{r}"].value = rr.branch_code
        cell_b = ws[f"B{r}"]
        cell_b.value = int(po_no)
        cell_b.number_format = "0"

        ws[f"C{r}"].value = (i + 1) * 10
        ws[f"D{r}"].value = rr.article
        ws[f"E{r}"].value = rr.description
        ws[f"F{r}"].value = po_str
        ws[f"G{r}"].value = deliv_str
        ws[f"H{r}"].value = None
        ws[f"I{r}"].value = float(rr.price_unit)
        ws[f"J{r}"].value = "EA"
        ws[f"K{r}"].value = float(rr.qty)
        ws[f"L{r}"].value = float(rr.price_unit * rr.qty)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

# ============================================================
# Streamlit UI
# ============================================================

st.set_page_config(page_title="PO Generator", layout="wide")
st.title("Online")

st.markdown(
"""
Upload:
- SPX / LEX / TT order files
- key.xlsx
"""
)


col1, col2 = st.columns(2)

with col1:
    key_file = st.file_uploader("Upload key.xlsx", type=["xlsx"])

with col2:
    po_date = st.date_input("PO date", value=dt.date.today())
    deliv_days = st.number_input("Delivery + days", min_value=0, max_value=365, value=7, step=1)

st.divider()

st.subheader("Upload platform files (optional: you can upload 1, 2, or all 3)")

spx_branch = st.text_input("SPX branch code (e.g., SPX903)", value="SPX903")
spx_file = st.file_uploader("Upload SPX file (.xlsx)", type=["xlsx"], key="spx")

lex_branch = st.text_input("LEX branch code (e.g., LEX903)", value="LEX903")
lex_file = st.file_uploader("Upload LEX file (.xlsx)", type=["xlsx"], key="lex")

tt_branch = st.text_input("TT branch code (e.g., TT905)", value="TT905")
tt_file = st.file_uploader("Upload TT file (.xlsx)", type=["xlsx"], key="tt")

st.divider()

# Validate + run
run = st.button("Generate PO", type="primary", use_container_width=True)

if run:
    try:
        if not key_file:
            st.error("Please upload key.xlsx")
            st.stop()

        inputs = []

        if spx_file:
            inputs.append(("SPX", spx_branch, spx_file.getvalue()))
        if lex_file:
            inputs.append(("LEX", lex_branch, lex_file.getvalue()))
        if tt_file:
            inputs.append(("TT", tt_branch, tt_file.getvalue()))

        if not inputs:
            st.error("Please upload at least one platform file")
            st.stop()

        rows = build_output_rows_from_uploads(
            key_bytes=key_file.getvalue(),
            inputs=inputs,
        )

        out_bytes = write_to_template_from_path(
            template_path=TEMPLATE_PATH,
            rows=rows,
            po_date=po_date,
            deliv_plus_days=int(deliv_days),
        )

        st.download_button(
            "Download Excel",
            data=out_bytes,
            file_name="output_filled.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(str(e))
